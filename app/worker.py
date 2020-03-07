import hashlib
import os
import re
from pathlib import Path
from time import sleep

import redis
import requests
from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import load_workbook

from app.core.celery_app import celery_app
from app.core.config import MIREA_SCHEDULE_URL
from app.db.models import Group, Period, ScheduleFile
from datetime import datetime


redis_client = redis.Redis(host='localhost', port=6379)


def md5(file_path):
    hash_md5 = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


# https://docs.celeryproject.org/en/latest/userguide/tasks.html#avoid-launching-synchronous-subtasks
# exceptions https://www.distributedpython.com/2018/09/28/celery-task-states/
@celery_app.task(acks_late=True)
def test_celery(word: str):
    sleep(5)
    return f"test task return {word}"


@celery_app.task(bind=True, track_started=True)
def download_files(self):
    folder = Path('schedule_files')
    os.makedirs(folder, exist_ok=True)

    session = requests.Session()
    page = session.get(MIREA_SCHEDULE_URL)
    soup = BeautifulSoup(page.text, "html.parser")
    links = soup.findAll('a', attrs={'href': re.compile(".+(IIT|IK)\s*.*\.(xlsx)$")})
    total = len(links)
    for i, link in enumerate(links):
        file_name = link.get('href').split('/')[-1]
        # logger.info(f'{i} {file_name}')
        response = session.get(link.get('href'), allow_redirects=True)
        temp_path = Path(folder / file_name)
        open(temp_path, 'wb').write(response.content)
        md5hash = md5(temp_path)
        hash_path = Path(folder / md5hash).with_suffix('.xlsx')
        os.rename(temp_path, hash_path)
        self.update_state(state='PROGRESS', meta={'done': i + 1, 'total': total})
        sleep(1)
    logger.info(self.AsyncResult(self.request.id).state)
    return 'finished'


def search_words(text, word_dict):
    for k, v in word_dict.items():
        for word in v:
            if re.search(rf'\w*{word}\w*', text, re.IGNORECASE):
                return k

    return 0


def parse_title(ws):
    course = category = grade = institute = 0
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=4):
        for cols in row:
            value = str(cols.value)
            if re.match(r"р\s*а\s*с\s*п\s*и\s*с\s*а\s*н\s*и\s*е\b", value, re.IGNORECASE):
                match1 = re.search(r'\w*\d\w*', value)
                if match1:
                    course = match1[0]

                category_dict = {
                    'class': ['занятий'],
                    'test': ['зачетной', 'зачётной', 'зачетов', 'зачётов'],
                    'exam': ['экзаменационной']
                }
                category = search_words(value, category_dict)

                match3 = re.search(r'\w*ИНТЕГУ\w*', value)
                if match3:
                    institute = "ИНТЕГУ"
                match3 = re.search(r'\w*КБиСП\w*', value, re.IGNORECASE) or re.search(r'\w*КБСП\w*', value,
                                                                                      re.IGNORECASE)
                if match3:
                    institute = "КБиСП"
                match3 = re.search(r'\w*кибернетики\w*', value)
                if match3:
                    institute = "ИК"
                match3 = re.search(r'\w*\bФизико\s*-\s*технологического\w*\b', value) or re.search(
                    r'\w*ФТИ\w*', value)
                if match3:
                    institute = "ФТИ"
                match3 = re.search(r'\w*\bИТ\s*\w*\b', value) or re.search(
                    r'\w*\bинформационных технологий\s*\w*\b', value)
                if match3:
                    institute = "ИТ"
                match3 = re.search(r'\w*РТС\w*', value)
                if match3:
                    institute = "РТС"
                match3 = re.search(r'\w*ИЭС\w*', value)
                if match3:
                    institute = "ИЭС"
                match3 = re.search(r'\w*ИЭП\w*', value)
                if match3:
                    institute = "ИЭП"
                match3 = re.search(r'\w*ВЗО\w*', value)
                if match3:
                    institute = "ИВЗО"
                match3 = re.search(r'\w*ИУСТРО\w*', value)
                if match3:
                    institute = "ИУСТРО"
                match3 = re.search(r'\w*ТХТ\w*', value) or re.search(r'\w*тонких химических технологий\w*',
                                                                     value)
                if match3:
                    institute = "ТХТ"
                grade = 'b'
                match4 = re.search(r'\w*магистратуры\w*', value)
                if match4:
                    grade = "m"

                result = {
                    'course': int(course),
                    'institute': institute,
                    'grade': grade,
                    'category': category
                }
                if 0 in result.values():
                    return 0
                else:
                    return result
    return 0


# async def insert_schedule_file(values):
#     return await ScheduleFile.objects.create(
#         year=2020,
#         semester=True,
#         institute=values.get("institute"),
#         course=values.get("course"),
#         grade=values.get("grade"),
#         category=values.get("category"),
#         file_name="54yyh",
#         updated=datetime.now()
#     )


@celery_app.task(bind=True, track_started=True)
def identify_files(self):
    folder = Path('schedule_files')
    path, dirs, files = next(os.walk(folder))
    file_count = len(files)
    sorted_files = sorted(os.listdir(folder))
    for i, file_name in enumerate(sorted_files):
        file_path = Path(folder / file_name)
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]

            result = parse_title(ws)

            if not result:
                logger.error(file_name)
            else:
                # logger.info(file_name)
                logger.info(result)
                # insert_schedule_file(result)

                break

    return 'finished'
